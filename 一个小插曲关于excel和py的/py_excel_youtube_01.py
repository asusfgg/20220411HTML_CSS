'''
Author: 千仞无锋
Date: 2022-04-19 20:56:46
LastEditors: 千仞无锋
LastEditTime: 2022-04-19 21:13:01
FilePath: \20220411HTML_CSS\一个小插曲关于excel和py的\py_excel_youtube_01.py
'''
import openpyxl
from openpyxl import Workbook,load_workbook
# openpyxl 支持2010版本office以上版本
# 读取文件
wb = load_workbook(
    r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
# 指定活跃工作表
ws = wb.active
print(ws['A1'].value)
