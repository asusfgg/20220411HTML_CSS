'''
Author: 千仞无锋
Date: 2022-04-20 21:31:30
LastEditors: 千仞无锋
LastEditTime: 2022-04-20 22:36:20
FilePath: \20220411HTML_CSS\一个小插曲关于excel和py的\youtube_openpyxl_03.py
'''
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font ,PatternFill, Border, Side, Alignment, Protection
# data 是一个列表，每个元素是一个字典
data = [{
    'name': '赵日天',
    'tall': 180,
    'age': 18,
    'weight': 80,
}, {
    'name': '钱多多',
    'tall': 190,
    'age': 19,
    'weight': 100
}, {
    'name': '孙小小',
    'tall': 210,
    'age': 22,
    'weight': 120
}]
# 写入excel
wb = Workbook()
ws = wb.active

title = ['姓名', '身高', '年龄', '体重']
ws.append(title)

for person in data:
    # 字典变列表
    ws.append(list(person.values()))
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\text.xlsx')

# 基础计算
# 调用公式
for col in range(2, 5):
    char = get_column_letter(col)
    ws[char + '7'] = f'=AVERAGE({char+"2"}:{char+"6"})'
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\text.xlsx')

# 字体
for col in range(1, 5):
    char = get_column_letter(col)
    ws[char + '1'].font = Font(bold=True,color='FF0000') # 字体加粗,变红
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\text.xlsx')
