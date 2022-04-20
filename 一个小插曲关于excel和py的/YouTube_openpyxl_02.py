'''
Author: fgg
Date: 2022-04-20 15:48:01
LastEditors: 千仞无锋
LastEditTime: 2022-04-20 21:30:15
FilePath: \20220411HTML_CSS\一个小插曲关于excel和py的\YouTube_openpyxl_02.py
Description: 学习用文件，主要就是笔记和随笔
Copyright (c) 2022 by fgg/Mechanical Design Studio, All Rights Reserved. 
'''
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
# 创建工作簿
# wb = Workbook()
# 激活工作表（默认的那个）
# ws = wb.active
# 设置工作表名称
# ws.title = 'qq'
# 填充单元格内容
# ws['A1'].value = 'Hello, openpyxl !'
# 哪怕是新建文件也要写全路径信息，我也不知道为啥，反正这么写就对了
# wb.save(r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
# 一次新增一横排的资料
# ws.append([123, 456, 789, 0])
# wb.save(r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')

# for i in range(5,10):
#     ws.append([111,222,333])
# wb.save(r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')

# 范围单元格资料
wb = load_workbook(
    r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
ws = wb.active
for row in range(2, 5):
    for col in range(2, 5):
        char = get_column_letter(col)
        ws[char+str(row)].value = '坐标是 {}*{}'.format(row, col)
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
for row in range(6, 10):
    for col in range(6, 10):
        char = get_column_letter(col)
        ws[char+str(row)].value = char+str(row)
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')

# 合并单元格
# ws.merge_cells('A1:E1')
# wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')

# # 恢复 但是内容没了
# ws.unmerge_cells('A1:E1')
# wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')

# 增加 行 ，或者 列
# 在第三横排做插入
# ws.insert_rows(3)
# ws['A3'].value = '这是新插入的row'
# wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
# # b列，不能写B 的写二
# col = column_index_from_string('B')
# ws.insert_cols(col)
# ws['B2'].value = '这是新插入的column'
# wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
# # 删除行列
# ws.delete_cols(col)
# wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')
ws.move_range('A1:E1', rows=10, cols=1)
wb.save(r'H:\study-notes\20220411HTML_CSS\一个小插曲关于excel和py的\new_excel.xlsx')