'''
Author: fgg
Date: 2022-04-20 11:07:47
LastEditors: fgg
LastEditTime: 2022-04-20 15:47:44
FilePath: \20220411HTML_CSS\一个小插曲关于excel和py的\YouTube_openpyxl_01.py
Description: 学习用文件，主要就是笔记和随笔
Copyright (c) 2022 by fgg/Mechanical Design Studio, All Rights Reserved. 
'''
from ctypes import wstring_at
import openpyxl
from openpyxl import workbook, load_workbook
wb = openpyxl.load_workbook(
    r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
ws = wb.active
print('-'*10+'打印原始A1单元格内容'+'-'*10)
print(ws['A1'].value)
ws['A1'].value = 'obj01'
print('-'*10+'打印修改后A1单元格内容'+'-'*10)
print(ws['A1'].value)
print('-'*10+'保存一下'+'-'*10)
# 此时修改成功了，但是文件没变，是因为只是修改，没有做储存
# 这里一定要写完整路径，我也不造为啥，但是这么写程序就能跑。。。
wb.save(r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
print('-'*10+'打印原始表单们的名称'+'-'*10)
print(wb.sheetnames)
# ws = wb['Sheet2']
# print('-'*30)
# 创建工作表
# 运行一次，多一个sheet，这不得了啊。。。写个if限制下
ws = wb.sheetnames
i = len(ws)
if i < 4:
    wb.create_sheet('qq')
    print('-'*10+'打印增加表单后的名单'+'-'*10)
    print(wb.sheetnames)
    print('-'*10+'保存一下'+'-'*10)
    wb.save(
        r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
else:
    print('-'*10+'打印增加表单后的名单'+'-'*10)
    print(wb.sheetnames)
    print('-'*10+'保存一下'+'-'*10)
    wb.save(
        r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
    exit()

