'''
Author: fgg
Date: 2022-04-19 09:26:07
LastEditors: 千仞无锋
LastEditTime: 2022-05-19 22:13:58
FilePath: \20220411HTML_CSS\一个小插曲关于excel和py的\py_excel_01.py
Description: 学习用文件，主要就是笔记和随笔
Copyright (c) 2022 by fgg/Mechanical Design Studio, All Rights Reserved. 
'''
# import the module of openpyxl
# 数字变字母，字母变数字
from openpyxl.utils import get_column_letter, column_index_from_string

import openpyxl
# open file
# wb means workbook
wb = openpyxl.load_workbook(
    r'C:\Users\Feng-DevWork\Desktop\20220411HTML_CSS\一个小插曲关于excel和py的\New.xlsx')
# show type of wb
# The console shows as : <class 'openpyxl.workbook.workbook.Workbook'>
print(type(wb))
# getting sheets from workbook
# the console shows as : ['Sheet1', 'Sheet2', 'Sheet3']
print(wb.sheetnames)
# get sheet by name
# iterate over sheet names
for sheet in wb:
#     # show on console the title name of sheet : sheet1 sheet2 sheet3
    print(sheet.title)
#     # create a new sheet
#     # note : This method will create a new sheet
mySheet = wb.create_sheet('MySheet')
# the console shows as : ['Sheet1', 'Sheet2', 'Sheet3', 'MySheet']
print(wb.sheetnames)

# behavioral function : Get the corresponding sheets by the variable name
# note : Kind of like how JavaScript gets elements ~~~~~
# note : function is deprecated
sheet3 = wb.get_sheet_by_name('ThirdSheet')
# fix
sheet3 = wb['Sheet3']
sheet1 = wb['Sheet1']
# the console shows as : <Worksheet "Sheet3">
print(sheet3)
# ws means worksheet
# .active means that this sheet was really active
ws = wb.active
# this is object of worksheet
print(type(ws))
# reference cell
# this is object of cell
print(ws['A1'])
# the content of A1 cell
# 这里的none 是因为 单元格有个底色，不认？？？别的就可以 , 拿表第一件事就是清除花里胡哨的东西！！！！！！！！！
print(ws['A1'].value)

c = ws['A4']
print('Row{},Column{} is{}'.format(c.row,c.column,c.value))
# coordinate means the position of cell 表示单元格的坐标位置
print('Cell {} is {}'.format(c.coordinate,c.value))
# reference of cells that is other methods
print(ws.cell(row=4,column=1))
print(ws.cell(row=4,column=1).value)

for i in range(1, 8, 2):
    第一行 到 第八行 步数为2
    print(i,ws.cell(row=i,column=1).value)

# getting rows and columns form the sheets
colC = ws['C']
print(colC)
print(colC[1])
print(colC[4].value)

row6 = ws[6]
# 切片
col_range = ws['A:C']
row_range = ws[1:4]
# 取出范围列里的每个单元格
for col in col_range:
     for cell in col:
        print(cell.value)
# 取出范围行里的每个单元格
for row in row_range:
    for cell in row:
        print(cell.value)
# 方法函数来访问单元格
for row in ws.iter_rows(min_row=1, max_row=2, max_col=2):
    for cell in row:
        print(cell.value)
#  tuple(要转换为元组的可迭代序列)返回元组
print(tuple(ws.rows))

# 单元格切片
# 指定单元格访问范围 先行后列输出
cell_range = ws['A1:C3']
for row_of_cell_obj in cell_range:
    for cell_obj in row_of_cell_obj:
#         # 值
        print(cell_obj.value)
#         # 坐标
        print(cell_obj.coordinate, cell_obj.value)
        print('-' * 20)
# # 输出几行几列
print('{}*{}'.format(ws.max_row, ws.max_column))
print('-'*30)
print(get_column_letter(2),get_column_letter(47),get_column_letter(900))
print('-'*30)
print(column_index_from_string('AA'),column_index_from_string('ZZ'),column_index_from_string('AAA'))
print('-'*30)

print('{}*{}'.format(ws.max_row, ws.max_column))
